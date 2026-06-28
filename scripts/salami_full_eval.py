#!/usr/bin/env python
"""
SALAMI Full Evaluation — all algorithms → Excel

Runs custom_librosa, foote, scluster, cnmf, fusion on every available
SALAMI song and writes scores to /app/data/salami_eval.xlsx.

Run inside worker-llm-1 (has MSAF + SALAMI data + worker code):

    docker exec worker-llm-1 python /app/workers/../scripts/salami_full_eval.py
    docker exec worker-llm-1 python /app/workers/../scripts/salami_full_eval.py --skip-cnmf
    docker exec worker-llm-1 python /app/workers/../scripts/salami_full_eval.py --max-tracks 20 --workers 4
"""
from __future__ import annotations

import argparse
import os
import re
import sys
import time
import traceback
from concurrent.futures import ThreadPoolExecutor, as_completed
from typing import Any

_here = os.path.dirname(os.path.abspath(__file__))
_app  = os.path.abspath(os.path.join(_here, ".."))
for p in (_app, _here):
    if p not in sys.path:
        sys.path.insert(0, p)

import numpy as np

from workers.segmenters.custom.segmentation_service import _analyze_content
from workers.segmenters.fusion.fusion_service import fuse_algorithm_results
from shared.segmentation.utils import (
    normalize_algorithm_result,
    boundaries_to_segments,
    normalize_boundaries,
    get_audio_duration,
)
from shared.labeling.heuristic import apply_two_layer_labels
from backend.services.salami_parser import parse_salami_annotation
from backend.services.evaluation_service import compute_boundary_metrics

ANNOTATIONS_DIR = "/app/data/salami/annotations"
AUDIO_CACHE     = "/app/data/audio_cache"
UPLOADS_DIR     = "/app/media/uploads"
OUT_XLSX        = "/app/data/salami_eval.xlsx"

ANNOTATORS = [1, 2]

ALGOS = ["custom_librosa", "foote", "scluster", "cnmf", "fusion"]

COLS = [
    "song_id", "title",
    "duration_s", "n_ref", "n_est",
    "precision", "recall", "f_measure",
    "annotator", "seg_time_s", "error",
]


# ── Audio discovery ───────────────────────────────────────────────────────────

def discover_audio() -> dict[str, str]:
    """Return {song_id: best_file_path} from cache + uploads, deduped by size."""
    best: dict[str, tuple[int, str]] = {}

    for name in os.listdir(AUDIO_CACHE) if os.path.isdir(AUDIO_CACHE) else []:
        m = re.match(r"^(\d+)\.mp3$", name)
        if m:
            p = os.path.join(AUDIO_CACHE, name)
            sid = m.group(1)
            sz = os.path.getsize(p)
            if sid not in best or sz > best[sid][0]:
                best[sid] = (sz, p)

    pat = re.compile(r"^[0-9a-f\-]+_(\d+)\.mp3$", re.IGNORECASE)
    for name in os.listdir(UPLOADS_DIR) if os.path.isdir(UPLOADS_DIR) else []:
        m = pat.match(name)
        if m:
            p = os.path.join(UPLOADS_DIR, name)
            sid = m.group(1)
            sz = os.path.getsize(p)
            if sid not in best or sz > best[sid][0]:
                best[sid] = (sz, p)

    # keep only songs that have annotations
    annotated = {
        d for d in os.listdir(ANNOTATIONS_DIR)
        if os.path.isdir(os.path.join(ANNOTATIONS_DIR, d)) and d.isdigit()
    } if os.path.isdir(ANNOTATIONS_DIR) else set()

    return {sid: path for sid, (_, path) in best.items() if sid in annotated}


# ── MSAF runner ───────────────────────────────────────────────────────────────

def _run_msaf(file_path: str, algorithm: str, task_id: str) -> dict:
    import msaf
    duration = get_audio_duration(file_path)
    est_times, _ = msaf.process(file_path, boundaries_id=algorithm)
    raw_times = [float(t) for t in est_times if t is not None] if est_times is not None else []
    boundaries = normalize_boundaries(raw_times, duration, min_gap_seconds=0.25, include_edges=True)
    segments = boundaries_to_segments(boundaries, duration, min_gap_seconds=0.1,
                                      boundary_metadata=[{"time": t, "source": algorithm,
                                                          "sources": [algorithm], "confidence": 1.0}
                                                         for t in boundaries])
    segments = apply_two_layer_labels(segments, file_path=file_path, duration_seconds=duration,
                                      method_hint="feature_clustering")
    return normalize_algorithm_result(
        task_id=task_id, status="completed", worker_type="msaf",
        algorithm=algorithm, duration_seconds=duration,
        boundaries=[{"time": t, "confidence": 1.0, "source": algorithm, "sources": [algorithm]}
                    for t in boundaries],
        segments=segments, diagnostics={},
    )


# ── Per-track evaluation ──────────────────────────────────────────────────────

def evaluate_track(song_id: str, file_path: str, algorithms: list[str]) -> dict[str, dict]:
    """Return {algo: row_dict} for one track."""
    task_id = f"eval_{song_id}"
    algo_results: dict[str, dict] = {}
    algo_times:   dict[str, float] = {}
    rows: dict[str, dict] = {}

    # 1. custom_librosa  (_analyze_content expects bytes, not a path)
    if "custom_librosa" in algorithms or "fusion" in algorithms:
        t0 = time.perf_counter()
        try:
            with open(file_path, "rb") as fh:
                audio_bytes = fh.read()
            result_raw = _analyze_content(audio_bytes, filename=os.path.basename(file_path))
            duration = get_audio_duration(file_path)
            segs = result_raw.get("segments", [])
            result = normalize_algorithm_result(
                task_id=task_id, status="completed", worker_type="custom",
                algorithm="custom_librosa", duration_seconds=duration,
                boundaries=[], segments=segs, diagnostics={},
            )
            algo_results["custom_librosa"] = result
        except Exception as e:
            algo_results["custom_librosa"] = {"status": "failed", "error": str(e), "segments": []}
        algo_times["custom_librosa"] = time.perf_counter() - t0

    # 2. MSAF algorithms
    for algo in ["foote", "scluster", "cnmf"]:
        if algo not in algorithms and "fusion" not in algorithms:
            continue
        t0 = time.perf_counter()
        try:
            algo_results[algo] = _run_msaf(file_path, algo, task_id)
        except Exception as e:
            algo_results[algo] = {"status": "failed", "error": str(e), "segments": []}
        algo_times[algo] = time.perf_counter() - t0

    # 3. fusion
    if "fusion" in algorithms:
        base = {k: v for k, v in algo_results.items()
                if k in ("custom_librosa", "foote", "scluster", "cnmf")
                and v.get("status") == "completed"}
        t0 = time.perf_counter()
        try:
            if len(base) >= 2:
                result = fuse_algorithm_results(base, task_id=task_id, file_path=file_path)
                algo_results["fusion"] = result
            else:
                algo_results["fusion"] = {"status": "failed", "error": "not enough base results", "segments": []}
        except Exception as e:
            algo_results["fusion"] = {"status": "failed", "error": str(e), "segments": []}
        algo_times["fusion"] = time.perf_counter() - t0

    # 4. Evaluate each requested algorithm against SALAMI annotations
    # parse_salami_annotation(song_id, annotator) → list[{start,end,label}] | None
    duration_s = round(get_audio_duration(file_path), 1)

    for algo in algorithms:
        result = algo_results.get(algo, {})
        segs = result.get("segments") or []
        error = None if result.get("status") == "completed" else result.get("error", "unknown error")

        row_base = {
            "song_id": song_id,
            "title": "",
            "duration_s": duration_s,
            "n_est": len(segs),
            "n_ref": 0,
            "precision": None,
            "recall": None,
            "f_measure": None,
            "annotator": "",
            "seg_time_s": round(algo_times.get(algo, 0.0), 2),
            "error": error,
        }

        if error or not segs:
            rows[algo] = row_base
            continue

        # Try both annotators, keep the best F-measure
        best_row = None
        for ann_idx in ANNOTATORS:
            ref_segs = parse_salami_annotation(song_id, ann_idx)
            if not ref_segs:
                continue
            try:
                metrics = compute_boundary_metrics(ref_segs, segs)
            except Exception:
                continue
            row = {**row_base,
                   "n_ref": len(ref_segs),
                   "precision": round(metrics.get("precision", 0.0), 4),
                   "recall":    round(metrics.get("recall", 0.0), 4),
                   "f_measure": round(metrics.get("f_measure", 0.0), 4),
                   "annotator": ann_idx}
            if best_row is None or (row["f_measure"] or 0) > (best_row["f_measure"] or 0):
                best_row = row
        rows[algo] = best_row or row_base

    return rows


# ── Excel writer ──────────────────────────────────────────────────────────────

def write_excel(all_rows: dict[str, list[dict]], out_path: str) -> None:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    wb.remove(wb.active)  # remove default sheet

    header_fill = PatternFill("solid", fgColor="1E293B")
    header_font = Font(bold=True, color="E2E8F0")
    good_font   = Font(color="22C55E", bold=True)
    mid_font    = Font(color="F59E0B", bold=True)
    bad_font    = Font(color="EF4444", bold=True)

    for algo, rows in all_rows.items():
        ws = wb.create_sheet(title=algo[:31])

        # Header
        ws.append(COLS)
        for col_idx, _ in enumerate(COLS, 1):
            cell = ws.cell(row=1, column=col_idx)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center")

        # Data rows
        for row in rows:
            ws.append([row.get(c) for c in COLS])
            # Colour f_measure
            f = row.get("f_measure")
            if f is not None:
                fm_col = COLS.index("f_measure") + 1
                cell = ws.cell(row=ws.max_row, column=fm_col)
                cell.font = good_font if f >= 0.5 else (mid_font if f >= 0.3 else bad_font)

        # Auto-width
        for col_idx, col_name in enumerate(COLS, 1):
            max_len = max((len(str(r.get(col_name) or "")) for r in rows), default=0)
            ws.column_dimensions[get_column_letter(col_idx)].width = max(len(col_name), max_len) + 2

    # Summary sheet
    ws_sum = wb.create_sheet(title="Summary", index=0)
    sum_cols = ["algorithm", "n_tracks", "macro_f1", "macro_precision", "macro_recall", "median_f1", "avg_seg_time_s"]
    ws_sum.append(sum_cols)
    for col_idx in range(1, len(sum_cols) + 1):
        cell = ws_sum.cell(row=1, column=col_idx)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")

    for algo, rows in all_rows.items():
        fms = [r["f_measure"] for r in rows if r.get("f_measure") is not None]
        prs = [r["precision"]  for r in rows if r.get("precision")  is not None]
        rcs = [r["recall"]     for r in rows if r.get("recall")     is not None]
        tms = [r["seg_time_s"] for r in rows if r.get("seg_time_s") is not None]
        n = len(rows)
        ws_sum.append([
            algo, n,
            round(float(np.mean(fms)), 4) if fms else None,
            round(float(np.mean(prs)), 4) if prs else None,
            round(float(np.mean(rcs)), 4) if rcs else None,
            round(float(np.median(fms)), 4) if fms else None,
            round(float(np.mean(tms)), 2) if tms else None,
        ])

    for col_idx, col_name in enumerate(sum_cols, 1):
        ws_sum.column_dimensions[get_column_letter(col_idx)].width = max(len(col_name) + 4, 16)

    wb.save(out_path)
    print(f"\nSaved: {out_path}")


# ── Main ──────────────────────────────────────────────────────────────────────

def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument("--algorithms", nargs="+",
                        default=["custom_librosa", "foote", "scluster", "fusion"],
                        help="Algorithms to evaluate. cnmf is slow — add explicitly if needed.")
    parser.add_argument("--max-tracks", type=int, default=0, help="0 = all")
    parser.add_argument("--workers",    type=int, default=3, help="Parallel tracks")
    parser.add_argument("--out",        default=OUT_XLSX)
    args = parser.parse_args()

    songs = discover_audio()
    if not songs:
        print("No annotated audio found. Check AUDIO_CACHE and UPLOADS_DIR.")
        sys.exit(1)

    if args.max_tracks > 0:
        songs = dict(list(songs.items())[: args.max_tracks])

    algorithms = args.algorithms
    print(f"Tracks:     {len(songs)}")
    print(f"Algorithms: {algorithms}")
    print(f"Workers:    {args.workers}")
    print(f"Output:     {args.out}")
    print()

    all_rows: dict[str, list[dict]] = {a: [] for a in algorithms}
    done = 0
    t_start = time.perf_counter()

    def process(item: tuple[str, str]) -> dict[str, dict]:
        sid, path = item
        return evaluate_track(sid, path, algorithms)

    with ThreadPoolExecutor(max_workers=args.workers) as pool:
        futures = {pool.submit(process, item): item[0] for item in songs.items()}
        for fut in as_completed(futures):
            sid = futures[fut]
            done += 1
            try:
                track_rows = fut.result()
                for algo, row in track_rows.items():
                    all_rows[algo].append(row)
                fms = [r["f_measure"] for r in track_rows.values() if r.get("f_measure") is not None]
                avg_f = f"avg_f1={np.mean(fms):.3f}" if fms else "no score"
                elapsed = time.perf_counter() - t_start
                eta = (elapsed / done) * (len(songs) - done)
                print(f"  [{done}/{len(songs)}] song={sid}  {avg_f}  elapsed={elapsed:.0f}s  ETA={eta:.0f}s")
            except Exception:
                print(f"  [{done}/{len(songs)}] song={sid}  ERROR")
                traceback.print_exc()
                for algo in algorithms:
                    all_rows[algo].append({"song_id": sid, "error": "exception", **{c: None for c in COLS if c not in ("song_id", "error")}})

    # Sort rows by song_id
    for algo in algorithms:
        all_rows[algo].sort(key=lambda r: int(r.get("song_id") or 0))

    write_excel(all_rows, args.out)

    # Print summary to stdout too
    print("\n── Summary ──────────────────────────────")
    for algo in algorithms:
        fms = [r["f_measure"] for r in all_rows[algo] if r.get("f_measure") is not None]
        if fms:
            print(f"  {algo:<20} n={len(fms):>3}  macro_f1={np.mean(fms):.4f}  median={np.median(fms):.4f}")
        else:
            print(f"  {algo:<20} no scores")
    print(f"\nTotal time: {time.perf_counter() - t_start:.0f}s")


if __name__ == "__main__":
    main()
